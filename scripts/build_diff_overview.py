"""Build an Excel overview of all rc_export2 diffs.

Sheet 1 — Summary: one row per object with counts per diff type.
Sheets 2..N — one sheet per object with the full diff detail.
"""
import re
from pathlib import Path

import pandas as pd
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_EXPORT = Path(__file__).parent.parent / "output" / "rc_export2"
_config = yaml.safe_load((Path(__file__).parent.parent / "config.yaml").read_text())
_SOURCE = _config["source_org"]
_TARGET = _config["target_org"]
_OUT_XLSX = _EXPORT / "diff_overview.xlsx"

# Object order matches the original list
OBJECTS = [
    "ProductSellingModel",
    "ProductRelationshipType",
    "ProductCatalog",
    "Pricing_Table__c",
    "Product2",
    "PriceAdjustmentSchedule",
    "ProductCategory",
    "ProductComponentGroup",
    "ProductClassification",
    "ProductSellingModelOption",
    "PricebookEntry",
    "ProductCategoryProduct",
    "ProductRelatedComponent",
    "ProductClassificationAttr",
    "PriceAdjustmentTier",
    "AttributeBasedAdjustment",
    "AttributeAdjustmentCondition",
    "ProductConfigurationFlow",
]

# Colour palette
_GREEN  = "C6EFCE"  # identical
_YELLOW = "FFEB9C"  # field diffs
_RED    = "FFC7CE"  # missing rows
_BLUE   = "BDD7EE"  # header
_GREY   = "D9D9D9"  # summary header


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


def _autofit(ws) -> None:
    """Set each column width to fit its content (capped at 80)."""
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 80)


def build_summary(diffs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    for obj in OBJECTS:
        df = diffs.get(obj, pd.DataFrame())
        src_csv = _EXPORT / f"{_SOURCE}_{obj}.csv"
        tgt_csv = _EXPORT / f"{_TARGET}_{obj}.csv"
        def _row_count(p):
            try:
                return len(pd.read_csv(p, encoding="utf-8-sig"))
            except Exception:
                return 0

        src_count = _row_count(src_csv) if src_csv.exists() else "?"
        tgt_count = _row_count(tgt_csv) if tgt_csv.exists() else "?"

        if df.empty or "_status" not in df.columns:
            rows.append({
                "Object": obj,
                f"{_SOURCE} rows": src_count,
                f"{_TARGET} rows": tgt_count,
                f"Only in {_SOURCE}": 0,
                f"Only in {_TARGET}": 0,
                "Field diffs": 0,
                "Identical": "yes",
            })
            continue

        statuses = df["_status"].astype(str)
        only_src  = (statuses == f"only_in_{_SOURCE}").sum()
        only_tgt  = (statuses == f"only_in_{_TARGET}").sum()
        field_diff = (statuses == "field_diff").sum()
        identical  = statuses.str.startswith("IDENTICAL").any()

        rows.append({
            "Object": obj,
            f"{_SOURCE} rows": src_count,
            f"{_TARGET} rows": tgt_count,
            f"Only in {_SOURCE}": int(only_src),
            f"Only in {_TARGET}": int(only_tgt),
            "Field diffs": int(field_diff),
            "Identical": "yes" if identical else "",
        })
    return pd.DataFrame(rows)


def style_summary(ws) -> None:
    # Header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill("2F5496")
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2):
        obj_val      = str(row[0].value or "")
        only_src     = int(row[3].value or 0)
        only_tgt     = int(row[4].value or 0)
        field_diff   = int(row[5].value or 0)
        identical    = str(row[6].value or "")

        if only_src or only_tgt or field_diff:
            fill = _fill(_RED) if (only_src or only_tgt) else _fill(_YELLOW)
        else:
            fill = _fill(_GREEN)

        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        row[0].alignment = Alignment(horizontal="left")  # object name left-aligned


def style_detail(ws) -> None:
    # Header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _fill("2F5496")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    status_col = None
    for i, cell in enumerate(ws[1], 1):
        if str(cell.value) == "_status":
            status_col = i
            break

    for row in ws.iter_rows(min_row=2):
        status = str(row[status_col - 1].value) if status_col else ""
        if status == f"only_in_{_SOURCE}":
            fill = _fill(_RED)
        elif status == f"only_in_{_TARGET}":
            fill = _fill(_RED)
        elif status == "field_diff":
            fill = _fill(_YELLOW)
        elif status.startswith("IDENTICAL"):
            fill = _fill(_GREEN)
        else:
            fill = _fill(_GREY)
        for cell in row:
            cell.fill = fill


def main():
    # Load all diff CSVs
    diffs: dict[str, pd.DataFrame] = {}
    for obj in OBJECTS:
        path = _EXPORT / f"diff_{obj}.csv"
        if path.exists():
            diffs[obj] = pd.read_csv(path, encoding="utf-8-sig", low_memory=False)

    summary_df = build_summary(diffs)

    with pd.ExcelWriter(_OUT_XLSX, engine="openpyxl") as writer:
        # Sheet 1: summary
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws_sum = writer.sheets["Summary"]
        style_summary(ws_sum)
        _autofit(ws_sum)

        # Sheets 2..N: per-object detail
        for obj in OBJECTS:
            df = diffs.get(obj, pd.DataFrame())
            if df.empty:
                df = pd.DataFrame([{"_status": "no diff data"}])
            # Truncate sheet name to Excel's 31-char limit
            sheet_name = obj[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            style_detail(ws)
            _autofit(ws)

    print(f"Written: {_OUT_XLSX}")


if __name__ == "__main__":
    main()
